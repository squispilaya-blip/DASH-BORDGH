# -*- coding: utf-8 -*-
"""
ui_comparador.py — Pestaña de comparación de dos reportes exportados.

El usuario sube dos archivos Excel del mismo RIS pero de fechas distintas
(exportados desde el dashboard), y se compara qué niños se vacunaron
y cuáles siguen faltando.
"""

import io
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from comparador_exportado import (
    DOSE_COLS, build_summary_df, compare_reportes, get_dose_cols_in,
    load_reporte, to_dataframe,
)


# ── Caché de carga ────────────────────────────────────────────────────────────

@st.cache_data(show_spinner="Cargando reporte...")
def _load(file_bytes: bytes) -> pd.DataFrame:
    return load_reporte(file_bytes)


# ── Export Excel ──────────────────────────────────────────────────────────────

def _export_excel(
    df_resumen: pd.DataFrame,
    df_detalle: pd.DataFrame,
    df_informe: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    _fill_ws(wb.active,          "Resumen por EESS",    df_resumen)
    _fill_ws(wb.create_sheet(),  "Detalle completo",    df_detalle)
    _fill_ws(wb.create_sheet(),  "Informe estadístico", df_informe, bold_last=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_ws(ws, title: str, df: pd.DataFrame, bold_last: bool = False) -> None:
    ws.title = title

    HEADER_FILL = PatternFill(start_color="1A6FA8", end_color="1A6FA8", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    TOTAL_FILL  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    TOTAL_FONT  = Font(bold=True)

    # Cabecera
    for ci, h in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    total_data_rows = len(df)

    # Datos
    for ri, row in enumerate(df.itertuples(index=False), 2):
        is_total = bold_last and (ri == total_data_rows + 1)
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL

    # Ancho de columnas
    for col in ws.columns:
        length = max(
            (max((len(ln) for ln in str(c.value or "").split("\n")), default=0)
             for c in col),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, 55)

    ws.freeze_panes = "A2"


# ── UI principal ──────────────────────────────────────────────────────────────

def render_comparison_tab() -> None:
    st.subheader("🔄 Comparar dos reportes del mismo RIS")
    st.caption(
        "Sube el reporte **antiguo** y el **nuevo** (exportados desde el dashboard). "
        "Identifica qué niños se vacunaron y quiénes siguen faltando."
    )

    # ── Carga de archivos ─────────────────────────────────────────────────────
    col_a, col_b = st.columns(2)
    with col_a:
        file_a = st.file_uploader(
            "📂 Reporte ANTIGUO (primera fecha)",
            type=["xlsx"],
            key="comp_a",
            help="El reporte de la fecha anterior (ej. primera semana de abril).",
        )
    with col_b:
        file_b = st.file_uploader(
            "📂 Reporte NUEVO (fecha más reciente)",
            type=["xlsx"],
            key="comp_b",
            help="El reporte más reciente (ej. primera semana de mayo).",
        )

    if not file_a or not file_b:
        st.info("📋 Carga ambos reportes Excel para ver la comparación.")
        return

    try:
        df_a = _load(file_a.read())
        df_b = _load(file_b.read())
    except Exception as e:
        st.error(f"Error al leer los archivos: {e}")
        return

    # ── Validación básica ─────────────────────────────────────────────────────
    if 'DNI' not in df_a.columns or 'DNI' not in df_b.columns:
        st.error(
            "Los archivos no tienen columna 'DNI'. "
            "Verifica que son reportes exportados del dashboard."
        )
        return

    ris_a = df_a['RIS'].iloc[0] if 'RIS' in df_a.columns else "—"
    ris_b = df_b['RIS'].iloc[0] if 'RIS' in df_b.columns else "—"

    st.success(
        f"**Reporte A:** {len(df_a):,} niños — RIS: {ris_a} &nbsp;|&nbsp; "
        f"**Reporte B:** {len(df_b):,} niños — RIS: {ris_b}"
    )

    # ── Comparación ───────────────────────────────────────────────────────────
    with st.spinner("Comparando reportes..."):
        results = compare_reportes(df_a, df_b)

    if not results:
        st.info("No se detectaron cambios entre los dos reportes.")
        return

    # ── Métricas globales ─────────────────────────────────────────────────────
    total_vac     = sum(1 for r in results if r['categoria'] in ('se_vacuno', 'parcial'))
    total_falta   = sum(1 for r in results if r['categoria'] == 'sigue_faltando')
    total_parcial = sum(1 for r in results if r['categoria'] == 'parcial')
    total_nuevos  = sum(1 for r in results if r['categoria'] == 'nuevo')

    m1, m2, m3, m4 = st.columns(4)
    m1.metric(
        "✅ Se vacunaron",
        total_vac,
        help="Niños que recibieron al menos una vacuna pendiente entre los dos reportes (incluye parciales)",
    )
    m2.metric(
        "❌ Siguen faltando",
        total_falta,
        help="Niños sin ninguna nueva vacuna que aún tienen dosis pendientes",
    )
    m3.metric(
        "⚠️ Parcial",
        total_parcial,
        help="Niños que se vacunaron en algunas dosis pero aún tienen otras pendientes",
    )
    m4.metric(
        "➕ Nuevos en padrón",
        total_nuevos,
        help="Niños presentes en el reporte nuevo pero no en el anterior",
    )

    st.divider()

    # ── Filtros ───────────────────────────────────────────────────────────────
    st.markdown("### 🔍 Filtros")

    all_dose_cols = [
        c for c in DOSE_COLS
        if c in set(get_dose_cols_in(df_a)) | set(get_dose_cols_in(df_b))
    ]

    f1, f2, f3 = st.columns(3)
    with f1:
        cat_opciones = {
            "Todos": None,
            "✅ Se vacunaron":    "se_vacuno",
            "❌ Siguen faltando": "sigue_faltando",
            "⚠️ Parcial":        "parcial",
        }
        cat_sel_label = st.selectbox("Mostrar", list(cat_opciones.keys()))
        cat_sel = cat_opciones[cat_sel_label]

    with f2:
        vacuna_sel = st.multiselect(
            "Filtrar por vacuna",
            options=all_dose_cols,
            help="Deja vacío para ver todas las vacunas del esquema.",
        )

    with f3:
        eess_options = sorted({
            r['EESS'] for r in results
            if r['EESS'] and r['EESS'] != 'nan'
        })
        eess_sel = st.multiselect("Filtrar por EESS", eess_options)

    # ── Tabla de resultados ───────────────────────────────────────────────────
    st.markdown("### 📋 Detalle nominal de niños")

    df_det = to_dataframe(results, dose_filter=vacuna_sel or None, cat_filter=cat_sel)

    if eess_sel:
        df_det = df_det[df_det['EESS'].isin(eess_sel)]

    if df_det.empty:
        st.warning("No hay registros con los filtros aplicados.")
    else:
        st.caption(f"Mostrando **{len(df_det):,}** registros")
        st.dataframe(
            df_det,
            use_container_width=True,
            hide_index=True,
            height=460,
            column_config={
                "RIS":                   st.column_config.TextColumn("RIS",                  width="medium"),
                "Zona Sanitaria":        st.column_config.TextColumn("Zona Sanitaria",       width="medium"),
                "EESS":                  st.column_config.TextColumn("EESS",                 width="medium"),
                "DNI":                   st.column_config.TextColumn("DNI",                  width="small"),
                "Nombres":               st.column_config.TextColumn("Nombres",              width="large"),
                "Prioridad actual":      st.column_config.TextColumn("Prioridad actual",     width="medium"),
                "Categoría":             st.column_config.TextColumn("Categoría",            width="medium"),
                "Vacunas administradas": st.column_config.TextColumn("Vacunas administradas",width="large"),
                "Vacunas pendientes":    st.column_config.TextColumn("Vacunas pendientes",   width="large"),
            },
        )

    # ── Resumen por EESS (en pantalla) ────────────────────────────────────────
    st.divider()
    st.markdown("### 🏥 Resumen por EESS")

    resumen_rows = []
    eess_groups: dict = {}
    for r in results:
        eess = r['EESS']
        if eess not in eess_groups:
            eess_groups[eess] = {'vac': 0, 'falta': 0, 'parcial': 0, 'nuevo': 0}
        cat = r['categoria']
        if cat == 'nuevo':
            eess_groups[eess]['nuevo'] += 1
        elif cat == 'parcial':
            eess_groups[eess]['parcial'] += 1
            eess_groups[eess]['vac']     += 1
        elif cat == 'se_vacuno':
            eess_groups[eess]['vac']   += 1
        elif cat == 'sigue_faltando':
            eess_groups[eess]['falta'] += 1

    for eess, s in sorted(eess_groups.items()):
        resumen_rows.append({
            'EESS':               eess,
            '✅ Se vacunaron':    s['vac'],
            '⚠️ Parcial':        s['parcial'],
            '❌ Siguen faltando': s['falta'],
            '➕ Nuevos':          s['nuevo'],
            'Total':              sum(s.values()),
        })

    df_resumen = pd.DataFrame(resumen_rows)

    if not df_resumen.empty:
        # Fila de totales en pantalla
        total_row_resumen = {
            'EESS':               'TOTAL',
            '✅ Se vacunaron':    df_resumen['✅ Se vacunaron'].sum(),
            '⚠️ Parcial':        df_resumen['⚠️ Parcial'].sum(),
            '❌ Siguen faltando': df_resumen['❌ Siguen faltando'].sum(),
            '➕ Nuevos':          df_resumen['➕ Nuevos'].sum(),
            'Total':              df_resumen['Total'].sum(),
        }
        df_resumen_display = pd.concat(
            [df_resumen, pd.DataFrame([total_row_resumen])], ignore_index=True
        )
        st.dataframe(df_resumen_display, use_container_width=True, hide_index=True)

    # ── Informe estadístico detallado (en pantalla) ───────────────────────────
    st.divider()
    st.markdown("### 📊 Informe estadístico detallado")
    st.caption(
        "Compara la cantidad de niños en cada reporte, "
        "vacunados en el periodo, nuevos ingresos y pendientes, por EESS y total RIS."
    )

    df_informe = build_summary_df(df_a, df_b, results)

    if not df_informe.empty:
        st.dataframe(
            df_informe,
            use_container_width=True,
            hide_index=True,
            column_config={
                "RIS":                         st.column_config.TextColumn("RIS",         width="medium"),
                "EESS":                        st.column_config.TextColumn("EESS",        width="large"),
                "Total Reporte A (1er corte)": st.column_config.NumberColumn("Reporte A", width="small",
                                               help="Niños en el primer reporte"),
                "Total Reporte B (2do corte)": st.column_config.NumberColumn("Reporte B", width="small",
                                               help="Niños en el segundo reporte"),
                "Vacunados en el periodo":     st.column_config.NumberColumn("Vacunados", width="small"),
                "Parcialmente vacunados":      st.column_config.NumberColumn("Parciales", width="small"),
                "Nuevos niños ingresados":     st.column_config.NumberColumn("Nuevos",    width="small"),
                "No vacunados en el periodo":  st.column_config.NumberColumn("No vac.",   width="small"),
                "Retirados del padrón":        st.column_config.NumberColumn("Retirados", width="small"),
            },
        )

    # ── Exportar ──────────────────────────────────────────────────────────────
    st.divider()
    col_btn, col_cap = st.columns([1, 3])
    with col_btn:
        if st.button("📥 Generar Excel comparación", use_container_width=True):
            with st.spinner("Generando Excel..."):
                df_all     = to_dataframe(results)          # sin filtros
                df_inf_exp = build_summary_df(df_a, df_b, results)
                df_res_exp = df_resumen if not df_resumen.empty else pd.DataFrame()
                excel_bytes = _export_excel(df_res_exp, df_all, df_inf_exp)
            fecha = date.today().strftime("%Y%m%d")
            st.download_button(
                label="⬇️ Descargar comparación",
                data=excel_bytes,
                file_name=f"comparacion_{ris_b}_{fecha}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    with col_cap:
        st.caption(
            "El Excel incluye **3 hojas**: "
            "'Resumen por EESS' · 'Detalle completo' · 'Informe estadístico' "
            f"con los {len(results):,} registros y totales por establecimiento y RIS."
        )
